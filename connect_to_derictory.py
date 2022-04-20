from openpyxl import load_workbook


def insert_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save('database.xlsx')

    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


def in_database(user: int) -> bool:
    '''
    возвращает True, если id пользователь есть в database
    '''
    for row in range(1, user_page.max_row + 1):
        if user == user_page.cell(row=row, column=1).value:
            return True
        return False


def insert_user(*args):     # принимает произвольное количество аргументов
    '''
    вносит нового пользователя в базу данных
    '''
    row = user_page.max_row + 1
    user_page.cell(row=row, column=1).value = args[0]
    user_page.cell(row=row, column=2).value = args[1]
    user_page.cell(row=row, column=3).value = args[2]
    user_page.cell(row=row, column=4).value = args[3]
    bd.save('database.xlsx')


bd = load_workbook('database.xlsx')
for sheet in bd:
    print(sheet.title)
stickers_page = bd['stickers']
user_page = bd['users']

stickers = {}
replies = {}

for row in range(1, stickers_page.max_row + 1):
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    reply_text = stickers_page.cell(row=row, column=3).value
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text
